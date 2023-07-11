import pandas as pd 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

#load data to dataframe
pd.set_option('display.max_columns', 500)
file_name = "/workspaces/1202/Datasets/video_game_sales.csv"
raw_df = pd.read_csv(file_name)

# 1.What country had the highest sales for N64?
N64_df = raw_df[raw_df['Platform'] == 'N64'] #filter by platform N64
N64_df = N64_df.groupby(['Platform'])[['NA_Sales','EU_Sales','JP_Sales']].sum().reset_index().T 
N64_df = N64_df.iloc[1:, :]
# .iloc -- selection using integer, [row_start:row_end, col_start:col_end]
# .loc -- selection using label
N64_df.columns = ['N64']
print(N64_df.head())

wb = Workbook()
ws1 = wb.create_sheet('Q1')

for r in dataframe_to_rows(N64_df, header=True, index=False):
    ws1.append(r)

ws2 = wb.create_sheet('Q2')
#2. What is the distribution of the genres for the top 100 games?
highest_sales_df = raw_df.sort_values(['Global_Sales'], ascending=False)
highest_sales_df = highest_sales_df.iloc[:100, :]

genre_count_df = highest_sales_df.value_counts(['Genre','Platform']).reset_index()
genre_count_df = genre_count_df.sort_values(['Genre'])

for r in dataframe_to_rows(genre_count_df, header=True, index=False):
    ws2.append(r)

wb.save('week9/week9_P2.xlsx')


